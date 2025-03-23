import pandas as pd
from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QLabel, QPushButton, 
    QFileDialog, QMessageBox, QTableView
)
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QStandardItemModel, QStandardItem
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
import os
import json
from loading_overlay import LoadingOverlay
import numpy as np

class SingleFileUploader(QWidget):
    """
    GUI component that handles uploading and processing settlement files.
    - Allows uploading a main AFC-triffi file
    - Allows uploading settlement reports from different payment apps
    - Provides functionality to generate summaries and merged documents
    """
    def __init__(self, app_names):
        super().__init__()
        
        # Load configuration for supported payment apps (PayTm, PhonePe, etc.)
        appdata_dir = os.getenv("APPDATA") if os.name == "nt" else os.path.expanduser("~/.config")
        app_folder = os.path.join(appdata_dir, "kochimetro")
        config_path = os.path.join(app_folder, "config.json")
        
        with open(config_path, "r") as file:
            self.config = json.load(file)
        
        self.app_names = list(self.config.keys())

        # Reuse styles from ExcelUploader
        drop_zone_style = """
            QLabel {
                background-color: white;
                border: 2px dashed #aaaaaa;
                border-radius: 8px;
                padding: 20px;
                color: #666666;
                font-size: 14px;
            }
            QLabel:hover {
                background-color: #f8f8f8;
                border-color: #666666;
            }
        """

        button_style = """
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                padding: 10px 20px;
                border-radius: 4px;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """

        # Apply styles and update layout
        self.main_layout = QVBoxLayout()
        self.main_layout.setSpacing(15)
        self.main_layout.setContentsMargins(20, 20, 20, 20)

        # File Upload Area for the single Excel file
        self.file_label = QLabel("Upload a single Excel file")
        self.file_label.setStyleSheet(drop_zone_style)
        self.file_label.setAlignment(Qt.AlignCenter)
        self.file_label.mousePressEvent = self.upload_file
        self.file_path = None
        self.main_layout.addWidget(self.file_label)

        # Unified Upload Area for Settlement Reports
        self.settlement_label = QLabel("Upload all settlement reports")
        self.settlement_label.setStyleSheet(drop_zone_style)
        self.settlement_label.setAlignment(Qt.AlignCenter)
        self.settlement_label.mousePressEvent = self.upload_settlement_files
        self.settlement_files = {}
        self.main_layout.addWidget(self.settlement_label)

        # File Table for the single file
        self.file_table = QTableView()
        self.main_layout.addWidget(self.file_table)

        # Get Merged Doc Button
        self.merged_doc_button = QPushButton("Get Merged Doc")
        self.merged_doc_button.clicked.connect(self.get_merged_doc)
        self.merged_doc_button.setStyleSheet(button_style)
        self.main_layout.addWidget(self.merged_doc_button)

        # Get Summary Button
        self.summary_button = QPushButton("Get Summary")
        self.summary_button.clicked.connect(self.get_summary)
        self.summary_button.setStyleSheet(button_style)
        self.main_layout.addWidget(self.summary_button)

        # Apply styles to widgets
        self.file_label.setStyleSheet(drop_zone_style)
        self.settlement_label.setStyleSheet(drop_zone_style)
        self.summary_button.setStyleSheet(button_style)

        # Update labels with icons
        self.file_label.setText("📄 AFC-triffi file here\n click to browse")
        self.settlement_label.setText("📊 App settlement reports here\n click to browse")
        self.summary_button.setText("Get Summary")

        self.setLayout(self.main_layout)

        # Add loading overlay
        self.loading_overlay = LoadingOverlay(self)

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self.loading_overlay.setFixedSize(self.size())

    def upload_file(self, event):
        self.file_path, _ = QFileDialog.getOpenFileName(self, "Select Excel File", "", "Excel Files (*.xlsx)")
        if self.file_path:
            self.file_label.setText(self.file_path.split('/')[-1])
            self.load_table(self.file_table, self.file_path)

    def upload_settlement_files(self, event):
        """
        Handles uploading of settlement report files from different payment apps.
        Validates that:
        1. Correct number of files are uploaded (one per payment app)
        2. Each file name contains the corresponding app name
        """
        files, _ = QFileDialog.getOpenFileNames(self, "Select Settlement Reports", "", "Excel Files (*.xlsx)")
        
        # Validate number of files matches number of payment apps
        if len(files) > len(self.app_names):
            QMessageBox.critical(self, "Error", f"More than the required number of files have been uploaded. Upload only {len(self.app_names)} files")
            return
        if len(files) < len(self.app_names):
            QMessageBox.critical(self, "Error", f"Less than the required number of files have been uploaded. Upload {len(self.app_names)} files")
            return

        # Map files to their corresponding apps based on filename
        if files:
            try:
                self.settlement_files = {self.get_app_name(os.path.basename(file)): file for file in files}
                self.settlement_label.setText(f"Uploaded {len(files)} settlement files")
            except Exception as e:
                QMessageBox.warning(self, "Error", f"Unrecognized file detected: {e.args[1]}. Make sure the name of the app exists in the filename")

    def get_app_name(self, file_name):
        app_keywords = self.app_names
        for keyword in app_keywords:
            if keyword.lower() in file_name.lower():
                return keyword
        raise Exception("Unrecognized file detected", file_name)
        

    def load_table(self, table_view, file_path):
        try:
            df = pd.read_excel(file_path)
            model = QStandardItemModel()

            # Set headers
            model.setHorizontalHeaderLabels(df.columns.tolist())

            # Populate data
            for row in df.itertuples(index=False):
                items = [QStandardItem(str(value)) for value in row]
                model.appendRow(items)

            table_view.setModel(model)

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error loading file:\n{str(e)}")

    def get_summary(self):
        if not self.file_path:
            QMessageBox.warning(self, "Error", "No file uploaded for the main file.")
            return

        if not self.settlement_files:
            QMessageBox.warning(self, "Error", "No settlement files uploaded.")
            return

        self.summary_button.setEnabled(False)
        self.merged_doc_button.setEnabled(False)
        
        self.loading_overlay.start_loading("Processing files...")

        try:
            df = pd.read_excel(self.file_path)
            self.loading_overlay.set_progress(20)

            original_df = df
            self.loading_overlay.set_progress(40)

            process = Process(original_df, self.settlement_files)
            self.loading_overlay.set_progress(70)

            save_path, _ = QFileDialog.getSaveFileName(self, "Save Summary File", "", "Excel Files (*.xlsx)")
            if save_path:
                if not save_path.endswith(".xlsx"):
                    save_path += ".xlsx"

                with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
                    if not process.sheet1.empty:
                        process.sheet1.to_excel(writer, sheet_name="Grouped Data", index=False)
                    else:
                        pd.DataFrame().to_excel(writer, sheet_name="No Data", index=False)

                self.loading_overlay.stop_loading()
                QMessageBox.information(self, "Success", f"Summary saved to:\n{save_path}")
            else:
                self.loading_overlay.stop_loading()
                QMessageBox.information(self, "Canceled", "Save operation was canceled.")

        except Exception as e:
            self.loading_overlay.stop_loading()
            QMessageBox.critical(self, "Error", f"Error processing file:\n{str(e)}")
        
        finally:
            # Re-enable buttons
            self.summary_button.setEnabled(True)
            self.merged_doc_button.setEnabled(True)

    def get_merged_doc(self):
        if not self.file_path:
            QMessageBox.warning(self, "Error", "No file uploaded for the main file.")
            return

        if not self.settlement_files:
            QMessageBox.warning(self, "Error", "No settlement files uploaded.")
            return

        self.summary_button.setEnabled(False)
        self.merged_doc_button.setEnabled(False)
        
        self.loading_overlay.start_loading("Processing files...")

        try:
            df = pd.read_excel(self.file_path)
            self.loading_overlay.set_progress(20)

            original_df = df
            self.loading_overlay.set_progress(40)

            process = Process(original_df, self.settlement_files)
            self.loading_overlay.set_progress(70)

            save_path, _ = QFileDialog.getSaveFileName(self, "Save Merged Document", "", "Excel Files (*.xlsx)")
            if save_path:
                if not save_path.endswith(".xlsx"):
                    save_path += ".xlsx"
                
                # Pre-filter the data before writing to Excel
                if not process.sheet4.empty:
                    # Add Action column efficiently using numpy
                    process.sheet4['Action'] = ''
                    
                    # Ensure comment_col is preserved
                    process.sheet4['comment_col'] = process.sheet4['comment_col'].fillna('No comment')
                    
                    # Write to Excel efficiently using a context manager
                    with pd.ExcelWriter(save_path, engine='openpyxl', mode='w') as writer:
                        # Write main sheet
                        process.sheet4.to_excel(writer, sheet_name="Merged Data", index=False)

                        # Find duplicates based on ticket numbers and include all occurrences
                        duplicate_mask = process.sheet4['TicketNUmber'].duplicated(keep=False)
                        duplicates = process.sheet4[duplicate_mask].sort_values('TicketNUmber')
                        # Filter out rows with empty or missing ticket numbers
                        duplicates = duplicates[duplicates['TicketNUmber'].notna() & (duplicates['TicketNUmber'] != 'MISSING')]
                        if not duplicates.empty:
                            duplicates.to_excel(writer, sheet_name="Duplicate Tickets", index=False)

                        # Create separate sheets for each ONDCapp
                        for app in process.sheet4['ONDCapp'].unique():
                            if pd.notna(app):  # Skip if app name is NaN
                                app_data = process.sheet4[process.sheet4['ONDCapp'] == app]
                                if not app_data.empty:
                                    app_data.to_excel(writer, sheet_name=f"{app} Data", index=False)
                        
                        # Write filtered sheets without creating separate DataFrames
                else:
                    pd.DataFrame().to_excel(save_path, sheet_name="No Data", index=False)

                # Add data validation more efficiently
                workbook = load_workbook(save_path)
                sheet = workbook["Merged Data"]
                action_col = sheet.max_column
                
                # Create validation rule once
                validation = DataValidation(
                    type="list",
                    formula1='"Option1,Option2,Option3"',
                    allow_blank=True
                )
                sheet.add_data_validation(validation)
                
                # Add validation to entire column range at once
                validation.add(f"{chr(64 + action_col)}2:{chr(64 + action_col)}{sheet.max_row}")
                
                workbook.save(save_path)
                self.loading_overlay.stop_loading()
                QMessageBox.information(self, "Success", f"Merged document saved to:\n{save_path}")
            else:
                self.loading_overlay.stop_loading()
                QMessageBox.information(self, "Canceled", "Save operation was canceled.")

        except Exception as e:
            self.loading_overlay.stop_loading()
            QMessageBox.critical(self, "Error", f"Error processing file:\n{str(e)}")
        
        finally:
            # Re-enable buttons
            self.summary_button.setEnabled(True)
            self.merged_doc_button.setEnabled(True)

class Process:
    """
    Core processing logic for settlement reconciliation.
    
    Flow:
    1. Loads and normalizes the main transaction data
    2. Loads and processes settlement files from different payment apps
    3. Merges settlement data with original transactions
    4. Identifies discrepancies (excess/shortage/settled)
    5. Generates summary reports
    """
    def __init__(self, original_df, settlement_files):
        self.original_df = original_df  # Main transaction data
        self.settlement_files = {}      # Settlement data from payment apps
        self.load_config()             # Load app-specific column mappings
        
        # Process each settlement file
        for app_name, file_path in settlement_files.items():
            app_name = app_name.lower()
            if app_name in self.app_mapping:
                self.settlement_files[app_name] = pd.read_excel(file_path)
        
        # Standardize formats and process data
        self._normalize_original_df()
        self._process_settlement_files()
        
        # Merge and analyze data
        self.merged_data = self._merge_settlement_data()
        self.grouped_data = self._summarize_transactions()
        
        # Prepare output sheets
        self.sheet1 = self.grouped_data    # Summary by app and date
        self.sheet4 = self.merged_data     # Detailed transaction matching

    def load_config(self):
        # Get AppData directory
        appdata_dir = os.getenv("APPDATA") if os.name == "nt" else os.path.expanduser("~/.config")
        app_folder = os.path.join(appdata_dir, "kochimetro")
        os.makedirs(app_folder, exist_ok=True)
        config_path = os.path.join(app_folder, "config.json")
        # Load or create config.json
        if os.path.exists(config_path):
            with open(config_path, "r") as file:
                self.app_mapping = json.load(file)
        else:
            # Default configuration
            default_config = {
                'easemytrip': {'id_col': 'TicketId', 'match_col': 'TicketNUmber', 'amount_col': 'TOTALAMOUNT', 'settle_col': 'SettlementAmount', 'date_col': 'Date', 'comment_col': 'TicketStatus'},
                'nammayathri': {'id_col': 'Ticket Id', 'match_col': 'TicketNUmber', 'amount_col': 'Total Amount', 'settle_col': 'Settlement Amount', 'date_col': 'Date', 'comment_col': 'Ticket Status'},
                'phonepe': {'id_col': 'Ticket Id', 'match_col': 'TicketNUmber', 'amount_col': 'TOTAL AMOUNT', 'settle_col': 'Settlement Amount', 'date_col': 'Date', 'comment_col': 'Ticket Status'},
                'paytm': {'id_col': 'Operator Reference Number', 'match_col': 'order_id', 'amount_col': 'Total Price', 'settle_col': 'Payable Amount', 'date_col': 'Settlement Date', 'comment_col': 'Payment Status'},
                'rapido': {'id_col': 'Network Order ID', 'match_col': 'transaction_ref_no', 'amount_col': 'TOTAL AMOUNT', 'settle_col': 'Settlement Amount', 'date_col': 'Date', 'comment_col': 'Ticket Status'},
                'redbus': {'id_col': 'Network Order ID(From ondcTxnId)', 'match_col': 'transaction_ref_no', 'amount_col': 'TOTAL AMOUNT', 'settle_col': 'Settlement Amount', 'date_col': 'Date', 'comment_col': 'Ticket Status'}
            }

            # Save the default configuration to the file
            with open(config_path, "w") as file:
                json.dump(default_config, file, indent=4)
            self.app_mapping = default_config

    def _normalize_original_df(self):
        self.original_df['ONDCapp'] = self.original_df['ONDCapp'].str.lower()
        self.original_df = self.original_df.replace('yathri', 'nammayathri')
        self.original_df['insertDT'] = self._standardize_date(self.original_df['insertDT'])

    def _process_settlement_files(self):
        for app_name, df in self.settlement_files.items():
            try:
                mapping = self.app_mapping[app_name]
                date_col = mapping['date_col']
                
                # Print debug info before processing
                print(f"\n[{app_name}] Date Processing:")
                print("Before normalization:")
                print(f"Column '{date_col}' - First 5 values:")
                print(df[date_col].head().to_string())
                print(f"Data type: {df[date_col].dtype}")
                
                # Handle UTC timestamps (like in nammayathri)
                if 'UTC' in str(df[date_col].iloc[0]):
                    df[date_col] = pd.to_datetime(df[date_col]).dt.tz_localize(None)
                else:
                    # Convert dates without timezone info
                    df[date_col] = pd.to_datetime(df[date_col], format='mixed')
                
                # Debug after datetime conversion
                print("\nAfter datetime conversion:")
                print(df[date_col].head().to_string())
                
                # Convert to string format (only date part)
                df[date_col] = df[date_col].dt.date.astype(str)
                
                # Debug after final formatting
                print("\nAfter final formatting:")
                print(df[date_col].head().to_string())
                print("----------------------------------------")
                
                self.settlement_files[app_name] = df
                
            except Exception as e:
                print(f"\n[{app_name}] Error processing dates:")
                print(f"  Error: {str(e)}")
                print(f"  Date column type: {type(df[date_col])}")
                print(f"  Raw date values: {df[date_col].head().tolist()}")
                continue

    def _merge_settlement_data(self):
        # Create base DataFrame with required columns
        merged_data = self.original_df[['insertDT', 'TicketNUmber', 'order_id', 
                                      'transaction_ref_no', 'ONDCapp', 'total_amount', 
                                      'QRCodePrice', 'booking_status', 'descCode', 'Remark']].copy()
        
        # Add empty settlement columns
        merged_data['amount_col'] = None
        merged_data['settle_col'] = None
        merged_data['comment_col'] = None
        merged_data['unsettled'] = None

        # Store original row count
        original_count = len(merged_data)
        print(f"Original row count: {original_count}")

        final_merged_data = pd.DataFrame()  

        # QUICK FIX: Force add comment column mappings to ensure they're used
        comment_cols = {
            'easemytrip': 'TicketStatus',
            'nammayathri': 'Ticket Status',
            'phonepe': 'Ticket Status',
            'paytm': 'Payment Status',
            'rapido': 'Ticket Status',
            'redbus': 'Ticket Status'
        }

        for app_name, mapping in self.app_mapping.items():
            try:
                # QUICK FIX: Add comment_col to mapping if not already there
                if 'comment_col' not in mapping and app_name in comment_cols:
                    mapping['comment_col'] = comment_cols[app_name]
                    print(f"Added comment_col '{mapping['comment_col']}' to {app_name} mapping")

                settlement_df = self.settlement_files.get(app_name)
                if settlement_df is None:
                    continue

                print(f"\n\n==== BEGIN PROCESSING {app_name.upper()} ====")
                
                # Get data for current app
                app_data = merged_data[merged_data['ONDCapp'] == app_name].copy()
                print(f"App data rows: {len(app_data)}")

                # Get required columns from settlement file
                required_cols = [
                    mapping['id_col'], 
                    mapping['amount_col'], 
                    mapping['settle_col'],
                    mapping['date_col']
                ]
                
                # Only add comment_col if it exists in the mapping
                comment_col_name = None
                if 'comment_col' in mapping:
                    required_cols.append(mapping['comment_col'])
                    comment_col_name = mapping['comment_col']
                    print(f"Comment column from config: '{comment_col_name}'")
                else:
                    print("No comment column in config for this app")
                
                # Debug settlement file columns to verify comment column exists
                print(f"Available columns in {app_name} settlement file: {settlement_df.columns.tolist()}")
                if comment_col_name:
                    if comment_col_name in settlement_df.columns:
                        print(f"FOUND: Comment column '{comment_col_name}' exists in settlement file")
                        print(f"  Sample values: {settlement_df[comment_col_name].head(3).tolist()}")
                        print(f"  Data type: {settlement_df[comment_col_name].dtype}")
                        print(f"  Contains NaN: {settlement_df[comment_col_name].isna().any()}")
                        
                        # Check for empty strings or whitespace
                        is_empty = (settlement_df[comment_col_name].astype(str).str.strip() == '')
                        print(f"  Contains empty strings: {is_empty.any()}")
                    else:
                        print(f"ERROR: Comment column '{comment_col_name}' NOT FOUND in settlement file")
                
                # Filter to only include columns that exist in the settlement_df
                valid_cols = [col for col in required_cols if col in settlement_df.columns]
                settlement_data = settlement_df[valid_cols].copy()
                print(f"Valid columns after filtering: {valid_cols}")

                # Create pre-merge dictionary for direct mapping approach
                comment_dict = {}
                if comment_col_name and comment_col_name in settlement_data.columns:
                    comment_dict = dict(zip(settlement_data[mapping['id_col']], 
                                          settlement_data[comment_col_name]))
                    print(f"Created comment dictionary with {len(comment_dict)} entries")
                    print(f"Sample dictionary entries: {list(comment_dict.items())[:3]}")

                # Handle duplicate id_col entries by aggregating amount and settlement columns
                if settlement_data[mapping['id_col']].duplicated().any():
                    print(f"Found duplicate IDs in settlement data - aggregating")
                    agg_dict = {
                        mapping['amount_col']: 'sum',
                        mapping['settle_col']: 'sum',
                        mapping['date_col']: 'first'
                    }
                    if 'comment_col' in mapping:
                        agg_dict[mapping['comment_col']] = 'first'
                    
                    settlement_data = settlement_data.groupby(mapping['id_col']).agg(agg_dict).reset_index()
                
                if app_name == 'nammayathri':
                    settlement_data[mapping['settle_col']] = settlement_data[mapping['settle_col']].clip(lower=0)
                
                if app_name == 'paytm':
                    settlement_data[mapping['id_col']] = settlement_data[mapping['id_col']].astype(str).str.replace('...', '').str.strip()

                print(f"Ready for merge: app_data={len(app_data)} rows, settlement_data={len(settlement_data)} rows")
                
                # Keep outer merge to get both unmatched original rows AND unmatched settlement rows
                merged = pd.merge(
                    app_data,
                    settlement_data,
                    left_on=mapping['match_col'],
                    right_on=mapping['id_col'],
                    how='outer',
                    suffixes=('', '_settlement')  # Important: Avoid column name conflicts
                )
               
                print(f"After merge: {len(merged)} rows")
                print(f"All columns after merge: {merged.columns.tolist()}")
                
                # Fill ONDCapp for settlement-only records
                merged['ONDCapp'] = merged['ONDCapp'].fillna(app_name)

                # Handle dates
                merged['insertDT'] = merged['insertDT'].fillna(merged[mapping['date_col']])

                # Fill all ID-related columns based on the mapping
                if mapping['match_col'] == 'TicketNUmber':
                    merged['TicketNUmber'] = merged['TicketNUmber'].fillna(merged[mapping['id_col']])
                elif mapping['match_col'] == 'order_id':
                    merged['order_id'] = merged['order_id'].fillna(merged[mapping['id_col']])
                elif mapping['match_col'] == 'transaction_ref_no':
                    merged['transaction_ref_no'] = merged['transaction_ref_no'].fillna(merged[mapping['id_col']])

                # Set settlement amounts
                merged['amount_col'] = merged[mapping['amount_col']]
                merged['settle_col'] = merged[mapping['settle_col']]
                
                # ATTEMPT 1: Direct approach using suffix pattern
                print("\n--- COMMENT COLUMN DEBUGGING ---")
                
                if comment_col_name:
                    comment_col_with_suffix = f"{comment_col_name}_settlement"
                    has_original = comment_col_name in merged.columns
                    has_with_suffix = comment_col_with_suffix in merged.columns
                    
                    print(f"Original column name: '{comment_col_name}' exists: {has_original}")
                    print(f"With suffix: '{comment_col_with_suffix}' exists: {has_with_suffix}")
                    
                    # Try three different approaches and print results:
                    
                    # 1. Try with column suffix
                    if has_with_suffix:
                        print("\nApproach 1: Using column with suffix")
                        merged['debug_comment1'] = merged[comment_col_with_suffix]
                        print(f"Sample values: {merged['debug_comment1'].head(3).tolist()}")
                    
                    # 2. Try with original column name
                    if has_original:
                        print("\nApproach 2: Using original column name")
                        merged['debug_comment2'] = merged[comment_col_name]
                        print(f"Sample values: {merged['debug_comment2'].head(3).tolist()}")
                    
                    # 3. Try with dictionary mapping
                    print("\nApproach 3: Using dictionary mapping")
                    merged['debug_comment3'] = merged[mapping['id_col']].map(comment_dict)
                    print(f"Sample values: {merged['debug_comment3'].head(3).tolist()}")
                    
                    # See which approach works best
                    print("\nFinal decision for comment column:")
                    if has_with_suffix:
                        merged['comment_col'] = merged[comment_col_with_suffix]
                        print(f"Using column '{comment_col_with_suffix}'")
                    elif has_original:
                        merged['comment_col'] = merged[comment_col_name]
                        print(f"Using column '{comment_col_name}'")
                    elif len(comment_dict) > 0:
                        merged['comment_col'] = merged[mapping['id_col']].map(comment_dict)
                        print("Using dictionary mapping approach")
                    else:
                        merged['comment_col'] = 'No comment'
                        print("No valid method found - using 'No comment'")
                else:
                    merged['comment_col'] = 'No comment'
                    print("No comment column in mapping - using 'No comment'")
                
                # Check final result
                print(f"\nFinal comment_col values: {merged['comment_col'].head(5).tolist()}")
                print(f"NaN values in comment_col: {merged['comment_col'].isna().sum()}")
                
                # Make sure all comment values are strings and replace NaN with 'No comment'
                merged['comment_col'] = merged['comment_col'].fillna('No comment')
                merged['comment_col'] = merged['comment_col'].astype(str)
                merged['comment_col'] = merged['comment_col'].replace('nan', 'No comment')
                merged['comment_col'] = merged['comment_col'].replace('None', 'No comment')
                
                print(f"After cleaning - final comment_col values: {merged['comment_col'].head(5).tolist()}")
                print(f"Final NaN values in comment_col: {merged['comment_col'].isna().sum()}")
                
                # Calculate unsettled amount (handle NaN values)
                if app_name in ['redbus', 'rapido']:
                    merged['unsettled'] = merged['QRCodePrice'].fillna(0) - merged['amount_col'].fillna(0)
                else:
                    merged['unsettled'] = merged['QRCodePrice'].fillna(0) - merged['settle_col'].fillna(0)
                
                # Cleanup debug columns before concatenation
                if 'debug_comment1' in merged.columns:
                    merged = merged.drop(columns=['debug_comment1'])
                if 'debug_comment2' in merged.columns:
                    merged = merged.drop(columns=['debug_comment2'])
                if 'debug_comment3' in merged.columns:
                    merged = merged.drop(columns=['debug_comment3'])

                # Add to final DataFrame - ensure we include all columns
                final_merged_data = pd.concat([
                    final_merged_data,
                    merged
                ])
                
                print(f"==== FINISHED PROCESSING {app_name.upper()} ====\n")

            except Exception as e:
                print(f"\n[{app_name}] Error: {str(e)}")
                import traceback
                traceback.print_exc()
                continue

        # Add unprocessed records from other apps
        processed_apps = set(self.settlement_files.keys())
        unprocessed_data = merged_data[~merged_data['ONDCapp'].isin(processed_apps)]
        final_merged_data = pd.concat([final_merged_data, unprocessed_data])

        # Print final statistics
        print(f"\nFINAL DATA STATISTICS:")
        print(f"Total rows: {len(final_merged_data)}")
        print(f"Comment column values:")
        print(final_merged_data['comment_col'].value_counts().head(10))
        print(f"NaN values in comment_col: {final_merged_data['comment_col'].isna().sum()}")

        # Add result column based on conditions
        conditions = [
            # Settled: All amounts present
            (final_merged_data['total_amount'].notna() & 
             final_merged_data['QRCodePrice'].notna() & 
             final_merged_data['settle_col'].notna() & 
             final_merged_data['amount_col'].notna()),
            
            # Shortage: Settlement amounts missing
            (final_merged_data['settle_col'].isna() & 
             final_merged_data['amount_col'].isna()),
            
            # Excess: Original amounts missing 
            (final_merged_data['total_amount'].isna() & 
             final_merged_data['QRCodePrice'].isna())
        ]

        choices = ['Settled', 'Shortage', 'Excess']
        final_merged_data['result'] = np.select(conditions, choices, default='Unknown')

        # Print summary of data
        final_count = len(final_merged_data)
        settlement_only = final_count - original_count
        print(f"\nData Summary:")
        print(f"Original rows: {original_count}")
        print(f"Final rows: {final_count}")
        print(f"Settlement-only rows: {settlement_only}")

        # Make sure to include the result column in the returned data
        columns = ['insertDT', 'TicketNUmber', 'order_id', 'transaction_ref_no', 'ONDCapp', 
            'total_amount', 'QRCodePrice', 'booking_status', 'descCode', 'Remark', 
            'amount_col', 'settle_col', 'unsettled', 'comment_col', 'result']
        
        # Filter columns that exist in the DataFrame to avoid KeyErrors
        existing_columns = [col for col in columns if col in final_merged_data.columns]
        return final_merged_data[existing_columns]

    def _summarize_transactions(self):
        grouped_data = self.merged_data.groupby(['ONDCapp', 'insertDT']).agg({
            'QRCodePrice': 'sum',
            'total_amount': 'sum',
            'amount_col': 'sum',
            'settle_col': 'sum',
            'comment_col': 'first'
        }).reset_index()
        grouped_data.rename(columns={
            'QRCodePrice': 'original_amount(afc)',
            'total_amount': 'original_amount(triffi)',
            'amount_col': 'total_amount',
            'settle_col': 'settlement_amount',
            'comment_col': 'comment'
        }, inplace=True)
        return grouped_data

    def _standardize_date(self, date_series):
        """Standardize dates to YYYY-MM-DD format"""
        try:
            # Convert to datetime, handling multiple formats
            if 'UTC' in str(date_series.iloc[0]):
                date_series = pd.to_datetime(date_series).dt.tz_localize(None)
            else:
                date_series = pd.to_datetime(date_series, format='mixed')
            
            # Convert to date only (removes time component)
            return date_series.dt.date.astype(str)
            
        except Exception as e:
            print(f"Error standardizing dates: {e}")
            return date_series
