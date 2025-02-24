import pandas as pd
from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, 
    QFileDialog, QMessageBox, QTableView
)
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QStandardItemModel, QStandardItem
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from loading_overlay import LoadingOverlay
import os
import numpy as np
import logging


class ExcelUploader(QWidget):
    def __init__(self):
        super().__init__()

        # Update the drop zone styling
        drop_zone_style = """
            QLabel {
                background-color: white;
                border: 2px dashed #aaaaaa;
                border-radius: 8px;
                padding: 20px;
                color: #666666;
                font-size: 14px;
            }
            QLabel:hover {
                background-color: #f8f8f8;
                border-color: #666666;
            }
        """

        # Update button styling
        button_style = """
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                padding: 10px 20px;
                border-radius: 4px;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
            QPushButton:pressed {
                background-color: #3d8b40;
            }
            QPushButton:disabled {
                background-color: #cccccc;
            }
        """

        # Update table styling
        table_style = """
            QTableView {
                background-color: white;
                border: 1px solid #cccccc;
                border-radius: 4px;
                gridline-color: #e0e0e0;
            }
            QHeaderView::section {
                background-color: #f5f5f5;
                padding: 6px;
                border: none;
                border-right: 1px solid #cccccc;
                border-bottom: 1px solid #cccccc;
            }
        """

        self.main_layout = QVBoxLayout()
        self.main_layout.setSpacing(15)  # Add spacing between widgets
        self.main_layout.setContentsMargins(20, 20, 20, 20)  # Add margins

        self.upload_layout = QHBoxLayout()
        # File 1 Upload Area
        self.file1_layout = QVBoxLayout()
        self.file1_label = QLabel("Upload AFC file")
        self.file1_label.setStyleSheet(drop_zone_style)
        self.file1_label.setAlignment(Qt.AlignCenter)
        self.file1_label.setAcceptDrops(True)
        self.file1_label.installEventFilter(self)
        self.file1_path = None
        self.file1_layout.addWidget(self.file1_label)

        self.file1_table = QTableView()
        self.file1_layout.addWidget(self.file1_table)
        self.upload_layout.addLayout(self.file1_layout)

        # File 2 Upload Area
        self.file2_layout = QVBoxLayout()
        self.file2_label = QLabel("Upload Triffy file")
        self.file2_label.setStyleSheet(drop_zone_style)
        self.file2_label.setAlignment(Qt.AlignCenter)
        self.file2_label.setAcceptDrops(True)
        self.file2_label.installEventFilter(self)
        self.file2_path = None
        self.file2_layout.addWidget(self.file2_label)

        self.file2_table = QTableView()
        self.file2_layout.addWidget(self.file2_table)
        self.upload_layout.addLayout(self.file2_layout)

        self.main_layout.addLayout(self.upload_layout)

        # Submit Button
        self.submit_button = QPushButton("Submit")
        self.submit_button.clicked.connect(self.submit)
        self.main_layout.addWidget(self.submit_button)

        # Apply styles
        self.file1_label.setStyleSheet(drop_zone_style)
        self.file2_label.setStyleSheet(drop_zone_style)
        self.submit_button.setStyleSheet(button_style)
        self.file1_table.setStyleSheet(table_style)
        self.file2_table.setStyleSheet(table_style)

        # Update labels with icons or better instructions
        self.file1_label.setText("📄 AFC file here\n click to browse")
        self.file2_label.setText("📄 Triffy file here\n click to browse")

        self.setLayout(self.main_layout)

        # Add loading overlay
        self.loading_overlay = LoadingOverlay(self)

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self.loading_overlay.setFixedSize(self.size())

    def eventFilter(self, source, event):
        if event.type() == event.DragEnter and source in [self.file1_label, self.file2_label]:
            if any(url.toLocalFile().endswith(".xlsx") for url in event.mimeData().urls()):
                event.accept()
            else:
                event.ignore()
            return True

        elif event.type() == event.Drop and source in [self.file1_label, self.file2_label]:
            file_path = event.mimeData().urls()[0].toLocalFile()
            if file_path.endswith(".xlsx"):
                if source == self.file1_label:
                    self.file1_path = file_path
                    self.file1_label.setText(f"File 1: {file_path}")
                    self.load_table(self.file1_table, self.file1_path)
                elif source == self.file2_label:
                    self.file2_path = file_path
                    self.file2_label.setText(f"File 2: {file_path}")
                    self.load_table(self.file2_table, self.file2_path)
            return True

        elif event.type() == event.MouseButtonPress and source in [self.file1_label, self.file2_label]:
            file_path, _ = QFileDialog.getOpenFileName(self, "Select Excel File", "", "Excel Files (*.xlsx)")
            if file_path:
                if source == self.file1_label:
                    self.file1_path = file_path
                    self.file1_label.setText(file_path.split('/')[-1])
                    self.load_table(self.file1_table, self.file1_path)
                elif source == self.file2_label:
                    self.file2_path = file_path
                    self.file2_label.setText(file_path.split('/')[-1])
                    self.load_table(self.file2_table, self.file2_path)
            return True

        return super().eventFilter(source, event)

    def load_table(self, table_view, file_path):
        try:
            # Start loading overlay
            self.loading_overlay.start_loading("Loading file...")
            
            # Read Excel file
            df = pd.read_excel(file_path)
            model = QStandardItemModel()

            # Set headers
            model.setHorizontalHeaderLabels(df.columns.tolist())

            # Populate data
            for row in df.itertuples(index=False):
                items = [QStandardItem(str(value)) for value in row]
                model.appendRow(items)

            table_view.setModel(model)

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error loading file:\n{str(e)}")
        finally:
            # Stop loading overlay
            self.loading_overlay.stop_loading()

    def submit(self):
        """Process the uploaded files"""
        try:
            # Validate file paths
            if not self.file1_path or not self.file2_path:
                QMessageBox.warning(self, "Error", "Both files must be uploaded.")
                return

            # Define paths
            # appdata_dir = os.getenv("APPDATA") if os.name == "nt" else os.path.expanduser("~/.config")
            # app_folder = os.path.join(appdata_dir, "kochimetro")
            # os.makedirs(app_folder, exist_ok=True)
            # do_not_delete_path = os.path.join(app_folder, "DO_NOT_DELETE.csv")

            # # Fallback to local directory if AppData is not writable
            # if not os.access(app_folder, os.W_OK):
            #     do_not_delete_path = os.path.join(os.path.dirname(__file__), "DO_NOT_DELETE.csv")
            #     os.makedirs(os.path.dirname(do_not_delete_path), exist_ok=True)

            # # Load existing data if available
            # existing_data = None
            # if os.path.exists(do_not_delete_path):
            #     try:
            #         existing_data = pd.read_csv(do_not_delete_path, low_memory=False)
            #     except Exception as e:
            #         logging.error(f"Error reading existing DO_NOT_DELETE.csv: {str(e)}")

            # Disable submit button and start loading
            self.submit_button.setEnabled(False)
            self.loading_overlay.start_loading("Processing files...")

            # Read and clean AFC data
            afc_df = pd.read_excel(self.file1_path)
            print("Original AFC sum:", afc_df['QRCodePrice'].sum())
            
            # Clean and validate AFC data with improved handling
            afc_df['TicketNUmber'] = afc_df['TicketNUmber'].astype(str).str.strip()
            afc_df['QRCodePrice'] = pd.to_numeric(afc_df['QRCodePrice'], errors='coerce')
            
            # Remove rows with null TicketNUmbers or QRCodePrices
            afc_df = afc_df.dropna(subset=['TicketNUmber', 'QRCodePrice'])
            
            # Log AFC data quality
            print("AFC Data Quality after cleaning:")
            print(f"Total rows: {len(afc_df)}")
            print(f"Unique TicketNUmbers: {afc_df['TicketNUmber'].nunique()}")
            print(f"Duplicate TicketNUmbers: {afc_df['TicketNUmber'].duplicated().sum()}")
            
            # Improved AFC aggregation logic
            def agg_desc_code(x):
                # Check for refunds first
                if any(str(code).upper() == 'REFUND' for code in x):
                    return 'REFUND'
                # If no refund, return the first non-null value
                valid_codes = [code for code in x if pd.notna(code)]
                return valid_codes[0] if valid_codes else 'UNKNOWN'

            # First, sort by insertDT to ensure chronological order
            afc_df = afc_df.sort_values('insertDT')
            
            # Aggregate AFC data with improved logic
            afc_df = afc_df.groupby('TicketNUmber', as_index=False).agg({
                'QRCodePrice': lambda x: x.sum(),  # Sum all prices
                'QRCodeId': 'first',  # Take the first QRCode ID
                'insertDT': 'last',   # Take the latest date
                'FromStation': 'first',
                'To Station': 'first',
                'ONDCapp': 'first',
                'descCode': agg_desc_code  # Use custom aggregation for descCode
            })
            
            print("AFC Data after aggregation:")
            print(f"Total unique tickets: {len(afc_df)}")
            print(f"AFC sum after aggregation: {afc_df['QRCodePrice'].sum():.2f}")
            
            self.loading_overlay.set_progress(30)

            # Read and clean Triffy data
            triffy_df = pd.read_excel(self.file2_path)
            triffy_df['ticket_number'] = triffy_df['ticket_number'].astype(str).str.strip()
            triffy_df['total_amount'] = pd.to_numeric(triffy_df['total_amount'], errors='coerce')
            self.loading_overlay.set_progress(60)

            # Aggregate AFC data with proper groupby
            afc_df = afc_df.groupby('TicketNUmber', as_index=False).agg({
                'QRCodePrice': 'sum',
                'QRCodeId': 'first',
                'insertDT': 'first',
                'FromStation': 'first',
                'To Station': 'first',
                'ONDCapp': 'first',
                'descCode': lambda x: 'REFUND' if 'REFUND' in x.values else x.iloc[0]
            })
            
            print("AFC sum after aggregation:", afc_df['QRCodePrice'].sum())

            # Aggregate Triffy data
            triffy_df = triffy_df.groupby('ticket_number', as_index=False).agg({
                'total_amount': 'sum',
                'transaction_ref_no': 'first',
                'order_id': 'first',
                'booking_status': 'first',
                'source': 'first',
                'destination': 'first',
                'booking_date': 'first'
            })

            rows_with_nan = triffy_df.isnull().any(axis=1).sum()
            logging.info(triffy_df.shape)
            logging.info(rows_with_nan)
            triffy_df = triffy_df.dropna()
            logging.info(triffy_df.shape)
            rows_with_nan = triffy_df.isnull().any(axis=1).sum()
            logging.info(rows_with_nan)

            # Merge with validation
            pre_merge_afc_sum = afc_df['QRCodePrice'].sum()
            merged_df = pd.merge(
                afc_df,
                triffy_df,
                left_on='TicketNUmber',
                right_on='ticket_number',
                how='outer',
                indicator=True
            )
            post_merge_afc_sum = merged_df['QRCodePrice'].sum()
            
            print(f"AFC sum before merge: {pre_merge_afc_sum}")
            print(f"AFC sum after merge: {post_merge_afc_sum}")

            # Convert dates
            merged_df['insertDT'] = pd.to_datetime(merged_df['insertDT']).dt.date
            merged_df['booking_date'] = pd.to_datetime(merged_df['booking_date']).dt.date

            # Simplified categorization logic
            merged_df['Remark'] = ''  # Initialize Remark column
            
            # Basic conditions
            is_refund = merged_df['descCode'].str.upper() == 'REFUND'
            amounts_match = np.isclose(merged_df['QRCodePrice'], merged_df['total_amount'], atol=0.01)
            full_refund = np.isclose(merged_df['QRCodePrice'], -merged_df['total_amount'], atol=0.01)
            triffi_revenue_more = merged_df['total_amount'] > merged_df['QRCodePrice']
            triffi_revenue_zero = merged_df['total_amount'] == 0
            afc_revenue_more = merged_df['QRCodePrice'] > merged_df['total_amount']

            # Categorize each record
            conditions = [
                (merged_df['_merge'] == 'left_only'),
                (merged_df['_merge'] == 'right_only'),
                (merged_df['_merge'] == 'both') & is_refund,
                (merged_df['_merge'] == 'both') & amounts_match,
                (merged_df['_merge'] == 'both') & full_refund,
                (merged_df['_merge'] == 'both') & triffi_revenue_zero,
                (merged_df['_merge'] == 'both') & afc_revenue_more,
                (merged_df['_merge'] == 'both') & triffi_revenue_more,
                (merged_df['_merge'] == 'both')  # catches any remaining 'both' cases
            ]
            
            choices = [
                'In AFC but not in Triffy',
                'In Triffy but not in AFC',
                'AFC Refund but Triffy Booked',
                'AFC = Triffy',
                'AFC Triffy Full Refund',
                'In AFC but not in Triffy',
                'AFC Revenue More than Triffy',
                'Triffy Revenue More than AFC',
                'Misc'
            ]
            
            merged_df['Remark'] = np.select(conditions, choices, default='Uncategorized')

            # Split into Errors and Equal sheets
            afc_equal_to_triffy = merged_df[merged_df['Remark'] == 'AFC = Triffy'].copy()
            final_df = merged_df[merged_df['Remark'] != 'AFC = Triffy'].copy()

            # Drop merge indicator column
            final_df.drop(columns=['_merge'], inplace=True, errors='ignore')
            afc_equal_to_triffy.drop(columns=['_merge'], inplace=True, errors='ignore')

            # Prepare final columns
            final_cols = [
                'TicketNUmber', 'QRCodeId', 'insertDT', 'FromStation', 'To Station',
                'total_amount', 'QRCodePrice', 'ONDCapp', 'transaction_ref_no',
                'order_id', 'booking_status', 'descCode', 'Remark'
            ]

            # Fill missing values appropriately
            numeric_cols = ['QRCodePrice', 'total_amount']
            other_cols = [col for col in final_cols if col not in numeric_cols]
            
            final_df[numeric_cols] = final_df[numeric_cols].fillna(0)
            final_df[other_cols] = final_df[other_cols].fillna("MISSING")
            afc_equal_to_triffy[numeric_cols] = afc_equal_to_triffy[numeric_cols].fillna(0)
            afc_equal_to_triffy[other_cols] = afc_equal_to_triffy[other_cols].fillna("MISSING")
            
            # Explicitly set column order for both DataFrames
            final_df = final_df[final_cols]
            afc_equal_to_triffy = afc_equal_to_triffy[final_cols]

            # Print column order before saving to verify
            print("Errors sheet columns:", list(final_df.columns))
            print("Equal sheet columns:", list(afc_equal_to_triffy.columns))

            # Save to Excel with optimized validation and column adjustment
            save_path, _ = QFileDialog.getSaveFileName(self, "Save Output File", "", "Excel Files (*.xlsx)")
            if save_path:
                if not save_path.endswith(".xlsx"):
                    save_path += ".xlsx"

                with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
                    # Add Action column while preserving order
                    final_df.insert(len(final_cols), 'Action', '')  # Add Action as the last column
                    afc_equal_to_triffy.insert(len(final_cols), 'Action', '')
                    
                    # Double-check column order after adding Action column
                    print("Final Errors columns:", list(final_df.columns))
                    print("Final Equal columns:", list(afc_equal_to_triffy.columns))
                    
                    final_df.to_excel(writer, sheet_name="Errors", index=False)
                    afc_equal_to_triffy.to_excel(writer, sheet_name="Equal", index=False)

                # Verify sums
                total_sum = final_df['QRCodePrice'].sum() + afc_equal_to_triffy['QRCodePrice'].sum()
                print(f"Sum of Errors sheet: {final_df['QRCodePrice'].sum()}")
                print(f"Sum of Equal sheet: {afc_equal_to_triffy['QRCodePrice'].sum()}")
                print(f"Total sum of Errors and Equal sheets: {total_sum}")

                if np.isclose(pre_merge_afc_sum, total_sum, rtol=1e-5):
                    print("Sums match: The total of Errors and Equal sheets equals the main QRCodePrice sum.")
                else:
                    print(f"Sums do not match: Main sum = {pre_merge_afc_sum}, Total of Errors and Equal = {total_sum}")
                    print(f"Difference: {pre_merge_afc_sum - total_sum}")

                # Log the results
                logging.info(f"Main QRCodePrice sum: {pre_merge_afc_sum}")
                logging.info(f"Sum of Errors sheet: {final_df['QRCodePrice'].sum()}")
                logging.info(f"Sum of Equal sheet: {afc_equal_to_triffy['QRCodePrice'].sum()}")
                logging.info(f"Total sum of Errors and Equal sheets: {total_sum}")
                if np.isclose(pre_merge_afc_sum, total_sum, rtol=1e-5):
                    logging.info("Sums match: The total of Errors and Equal sheets equals the main QRCodePrice sum.")
                else:
                    logging.warning(f"Sums do not match: Main sum = {pre_merge_afc_sum}, Total of Errors and Equal = {total_sum}")
                    logging.warning(f"Difference: {pre_merge_afc_sum - total_sum}")

                # Add dropdown validation
                workbook = load_workbook(save_path)
                action_options = ["Option 1", "Option 2", "Option 3", "Option 4", "Option 5"]
                
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    
                    # Create data validation for Action column once
                    dv = DataValidation(
                        type="list",
                        formula1=f'"{",".join(action_options)}"',
                        allow_blank=True
                    )
                    sheet.add_data_validation(dv)
                    
                    # Get the Action column letter (last column)
                    action_column = sheet.cell(1, sheet.max_column).column_letter
                    
                    # Apply validation to entire column range at once
                    dv.add(f'{action_column}2:{action_column}{sheet.max_row}')
                    
                    # Optimize column width adjustment
                    for column in sheet.columns:
                        # Sample only first 1000 rows for width calculation
                        sample_length = min(1000, sheet.max_row)
                        max_length = max(
                            len(str(cell.value or "")) 
                            for cell in column[:sample_length] 
                            if cell.value is not None
                        )
                        # Set a reasonable maximum width
                        adjusted_width = min(max_length + 2, 50)
                        sheet.column_dimensions[column[0].column_letter].width = adjusted_width

                workbook.save(save_path)
                self.loading_overlay.stop_loading()
                
                # Show success message with better formatting
                msg = QMessageBox()
                msg.setIcon(QMessageBox.Information)
                msg.setWindowTitle("Success")
                msg.setText("Files processed successfully!")
                msg.setInformativeText(f"Output saved to:\n{save_path}")
                msg.exec_()
            else:
                self.loading_overlay.stop_loading()
                QMessageBox.information(self, "Canceled", "Save operation was canceled.")

        except Exception as e:
            self.loading_overlay.stop_loading()
            logging.error(f"Error processing files: {str(e)}")
            QMessageBox.critical(self, "Error", f"Error processing files:\n{str(e)}")

        finally:
            self.submit_button.setEnabled(True)